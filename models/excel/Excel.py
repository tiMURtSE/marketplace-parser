import openpyxl
import os

from typing import Union

class Excel:
    
    def __init__(self, file_path: str, sheet: Union[str, None] = None):
        self._file_path = file_path
        self._book = openpyxl.load_workbook(file_path)
        self._sheet = self._book[sheet] if sheet else self._book.active
        self._show_info()

    def get_columns_data(self, columns: list[str], start_pos: int = 2, end_pos: int = 9999):
        columns_data = []
        column_indexes = self._get_all_column_indexes(columns)

        for row in self._sheet.iter_rows(min_row=start_pos, max_row=end_pos, values_only=True):
            # Получение значения каждого столбца
            row_data = [row[index - 1] for index in column_indexes]

            # Если значения столбцов в текущей строке пустые, то выйти из цикла
            if not any(row_data):
                break

            columns_data.append(row_data)

        return columns_data
    
    def write_data(self, columns: list[str], data: list, start_pos: int = 2):
        column_indexes = self._get_all_column_indexes(columns)

        if not self._is_start_pos_empty(start_pos, column_indexes):
            raise Exception("Стартовая позиция для заполнения данными занята")

        for row_index, row in enumerate(data):
            for index, column_index in enumerate(column_indexes):
                if self._check_column_equality(columns, row):
                    value = row[index]
                    self._sheet.cell(row=start_pos + row_index, column=column_index, value=value)
                else:
                    raise IndexError("Количество столбцов должно равняться с количеством значений, записываемых в эти столбцы")

        self._save()

    def _get_column_index(self, column_title: str):
        for col in self._sheet.iter_cols(min_row=1, max_row=1, min_col=1):
            col_value = col[0].value
            col_index = col[0].column

            if col_value == column_title:
                return col_index
            
        raise Exception(f"Столбец с названием {column_title} не найден")
            
    def _get_all_column_indexes(self, columns: list[str]):
        column_indexes = []

        for column in columns:
            if isinstance(column, str):
                column_indexes.append(self._get_column_index(column))
            else:
                raise Exception("Введены неверные данные столбцов")
            
        return column_indexes

    # def _format_row_data(self, row_data: list[str], columns: list[str | int]):
    #     """
    #     Добавление к данным из столбцов названия этих столбцов в виде ключа объекта. `["red", "AR-32"] -> { "color": "red", "article": "AR-32" }`
    #     """
    #     formatted_data = {}

    #     for index, data in enumerate(row_data):
    #         if isinstance(columns[index], str):
    #             formatted_data[columns[index]] = data
    #         elif isinstance(columns[index], int):
    #             return row_data
    #         else:
    #             raise Exception("Введены неверные данные столбцов")

    #     return formatted_data

    def _show_info(self):
        filename = os.path.basename(self._file_path)
        sheet_name = self._sheet.title

        print(f"Открыт файл: {filename} c листом \"{sheet_name}\"")

    def _check_column_equality(self, columns: list[str], data: list):
        return len(columns) == len(data)
    
    def _is_start_pos_empty(self, start_pos: int, column_indexes: list[int]):
        for index in column_indexes:
            cell_value = self._sheet.cell(row=start_pos, column=index).value

            if cell_value:
                return False

        return True

    def _save(self):
        self._book.save(self._file_path)